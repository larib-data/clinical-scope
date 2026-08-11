"""Tests for database_options_xlsx converter."""

import io
from pathlib import Path

import openpyxl
import pytest

from clinical_scope.database_options_xlsx import xlsx_bytes_to_database_options

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _build_xlsx(
    signals_rows: list[list],
    loops_rows: list[list] | None = None,
    spectrograms_rows: list[list] | None = None,
    psds_rows: list[list] | None = None,
) -> bytes:
    """
    Build a minimal XLSX bytes object with a ``signals`` sheet and optional other sheets.

    *signals_rows* must include the header as the first element.
    The optional sheets' rows must likewise include their header as the first element.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "signals"
    for row in signals_rows:
        ws.append(row)

    if loops_rows is not None:
        ws_loops = wb.create_sheet("loops")
        for row in loops_rows:
            ws_loops.append(row)

    if spectrograms_rows is not None:
        ws_spectrograms = wb.create_sheet("spectrograms")
        for row in spectrograms_rows:
            ws_spectrograms.append(row)

    if psds_rows is not None:
        ws_psds = wb.create_sheet("psds")
        for row in psds_rows:
            ws_psds.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


SIGNALS_HEADER = [
    "datasource",
    "signal",
    "label",
    "unit",
    "unit_conversion",
    "range_min",
    "range_max",
    "priority",
    "color",
    "visible",
    "line_dash",
    "period_resampling",
    "display",
    "groups",
]

LOOPS_HEADER = ["datasource", "loop_name", "x_signal", "y_signal"]

SPECTROGRAMS_HEADER = [
    "datasource",
    "spectrogram_name",
    "signal",
    "freq_min",
    "freq_max",
    "db_min",
    "db_max",
    "window_s",
    "overlap",
]

PSDS_HEADER = [
    "datasource",
    "groups",
    "signal",
    "freq_min",
    "freq_max",
    "db_min",
    "db_max",
    "window_s",
    "overlap",
    "label",
    "color",
    "line_dash",
]


# ---------------------------------------------------------------------------
# Tests: basic signal parsing
# ---------------------------------------------------------------------------


class TestBasicSignalParsing:
    def test_single_signal_minimal(self):
        data = _build_xlsx(
            [
                SIGNALS_HEADER,
                ["ds_a", "SIG1", "", "", "", "", "", "", "", "", "", "", "", ""],
            ]
        )
        result = xlsx_bytes_to_database_options(data)
        assert "ds_a" in result
        assert "SIG1" in result["ds_a"]["signals"]

    def test_signal_label_included_when_different_from_name(self):
        data = _build_xlsx(
            [
                SIGNALS_HEADER,
                ["ds_a", "SIG1", "My label", "", "", "", "", "", "", "", "", "", "", ""],
            ]
        )
        result = xlsx_bytes_to_database_options(data)
        assert result["ds_a"]["signals"]["SIG1"]["label"] == "My label"

    def test_signal_label_omitted_when_equal_to_name(self):
        data = _build_xlsx(
            [
                SIGNALS_HEADER,
                ["ds_a", "SIG1", "SIG1", "", "", "", "", "", "", "", "", "", "", ""],
            ]
        )
        result = xlsx_bytes_to_database_options(data)
        assert "label" not in result["ds_a"]["signals"]["SIG1"]

    def test_unit_and_unit_conversion(self):
        data = _build_xlsx(
            [
                SIGNALS_HEADER,
                ["ds_a", "SIG1", "", "mmHg", "1.35951", "", "", "", "", "", "", "", "", ""],
            ]
        )
        sig = xlsx_bytes_to_database_options(data)["ds_a"]["signals"]["SIG1"]
        assert sig["unit"] == "mmHg"
        assert sig["unit_conversion"] == pytest.approx(1.35951)

    def test_range_both_bounds(self):
        data = _build_xlsx(
            [
                SIGNALS_HEADER,
                ["ds_a", "SIG1", "", "", "", "-5", "25", "", "", "", "", "", "", ""],
            ]
        )
        sig = xlsx_bytes_to_database_options(data)["ds_a"]["signals"]["SIG1"]
        assert sig["range"] == [-5.0, 25.0]

    def test_range_only_min(self):
        data = _build_xlsx(
            [
                SIGNALS_HEADER,
                ["ds_a", "SIG1", "", "", "", "-5", "", "", "", "", "", "", "", ""],
            ]
        )
        sig = xlsx_bytes_to_database_options(data)["ds_a"]["signals"]["SIG1"]
        assert sig["range"] == [-5.0, None]

    def test_range_absent_when_empty(self):
        data = _build_xlsx(
            [
                SIGNALS_HEADER,
                ["ds_a", "SIG1", "", "", "", "", "", "", "", "", "", "", "", ""],
            ]
        )
        sig = xlsx_bytes_to_database_options(data)["ds_a"]["signals"]["SIG1"]
        assert "range" not in sig

    def test_color_priority_line_dash(self):
        data = _build_xlsx(
            [
                SIGNALS_HEADER,
                ["ds_a", "SIG1", "", "", "", "", "", "2", "red", "", "dash", "", "", ""],
            ]
        )
        sig = xlsx_bytes_to_database_options(data)["ds_a"]["signals"]["SIG1"]
        assert sig["color"] == "red"
        assert sig["priority"] == pytest.approx(2.0)
        assert sig["line_dash"] == "dash"

    def test_period_resampling_on_signal(self):
        data = _build_xlsx(
            [
                SIGNALS_HEADER,
                ["ds_a", "SIG1", "", "", "", "", "", "", "", "", "", "0.8", "", ""],
            ]
        )
        sig = xlsx_bytes_to_database_options(data)["ds_a"]["signals"]["SIG1"]
        assert sig["period_resampling"] == pytest.approx(0.8)


# ---------------------------------------------------------------------------
# Tests: display column → field_display
# ---------------------------------------------------------------------------


class TestDisplayColumn:
    def test_empty_display_means_visible(self):
        data = _build_xlsx(
            [
                SIGNALS_HEADER,
                ["ds_a", "SIG1", "", "", "", "", "", "", "", "", "", "", "", ""],
            ]
        )
        result = xlsx_bytes_to_database_options(data)
        assert "SIG1" in result["ds_a"]["field_display"]

    def test_display_yes_explicit(self):
        data = _build_xlsx(
            [
                SIGNALS_HEADER,
                ["ds_a", "SIG1", "", "", "", "", "", "", "", "", "", "", "yes", ""],
            ]
        )
        result = xlsx_bytes_to_database_options(data)
        assert "SIG1" in result["ds_a"]["field_display"]

    def test_display_no_excludes_from_field_display(self):
        data = _build_xlsx(
            [
                SIGNALS_HEADER,
                ["ds_a", "SIG1", "", "", "", "", "", "", "", "", "", "", "no", ""],
                ["ds_a", "SIG2", "", "", "", "", "", "", "", "", "", "", "yes", ""],
            ]
        )
        result = xlsx_bytes_to_database_options(data)
        assert "SIG1" not in result["ds_a"]["field_display"]
        assert "SIG2" in result["ds_a"]["field_display"]

    def test_display_zero_excludes(self):
        data = _build_xlsx(
            [
                SIGNALS_HEADER,
                ["ds_a", "SIG1", "", "", "", "", "", "", "", "", "", "", "0", ""],
            ]
        )
        result = xlsx_bytes_to_database_options(data)
        assert "SIG1" not in result["ds_a"].get("field_display", [])


# ---------------------------------------------------------------------------
# Tests: visible column
# ---------------------------------------------------------------------------


class TestVisibleColumn:
    def test_empty_visible_not_added_to_sig_opts(self):
        data = _build_xlsx(
            [
                SIGNALS_HEADER,
                ["ds_a", "SIG1", "", "", "", "", "", "", "", "", "", "", "", ""],
            ]
        )
        sig = xlsx_bytes_to_database_options(data)["ds_a"]["signals"]["SIG1"]
        assert "visible" not in sig

    def test_visible_no_adds_false(self):
        data = _build_xlsx(
            [
                SIGNALS_HEADER,
                ["ds_a", "SIG1", "", "", "", "", "", "", "", "no", "", "", "", ""],
            ]
        )
        sig = xlsx_bytes_to_database_options(data)["ds_a"]["signals"]["SIG1"]
        assert sig["visible"] is False

    def test_visible_yes_not_added(self):
        data = _build_xlsx(
            [
                SIGNALS_HEADER,
                ["ds_a", "SIG1", "", "", "", "", "", "", "", "yes", "", "", "", ""],
            ]
        )
        sig = xlsx_bytes_to_database_options(data)["ds_a"]["signals"]["SIG1"]
        assert "visible" not in sig


# ---------------------------------------------------------------------------
# Tests: sentinel row (*) → numerics
# ---------------------------------------------------------------------------


class TestSentinelRow:
    def test_sentinel_creates_numerics(self):
        data = _build_xlsx(
            [
                SIGNALS_HEADER,
                ["ds_a", "*", "", "", "", "", "", "2.5", "", "", "", "0.2", "", ""],
            ]
        )
        result = xlsx_bytes_to_database_options(data)
        assert result["ds_a"]["numerics"] == {"priority": 2.5, "period_resampling": 0.2}

    def test_sentinel_partial_numerics(self):
        data = _build_xlsx(
            [
                SIGNALS_HEADER,
                ["ds_a", "*", "", "", "", "", "", "3", "", "", "", "", "", ""],
            ]
        )
        result = xlsx_bytes_to_database_options(data)
        assert result["ds_a"]["numerics"] == {"priority": 3.0}

    def test_sentinel_not_added_to_signals(self):
        data = _build_xlsx(
            [
                SIGNALS_HEADER,
                ["ds_a", "*", "", "", "", "", "", "1", "", "", "", "0.5", "", ""],
                ["ds_a", "SIG1", "", "", "", "", "", "", "", "", "", "", "", ""],
            ]
        )
        result = xlsx_bytes_to_database_options(data)
        assert "*" not in result["ds_a"].get("signals", {})

    def test_sentinel_not_added_to_field_display(self):
        data = _build_xlsx(
            [
                SIGNALS_HEADER,
                ["ds_a", "*", "", "", "", "", "", "1", "", "", "", "", "", ""],
            ]
        )
        result = xlsx_bytes_to_database_options(data)
        assert "*" not in result["ds_a"].get("field_display", [])

    def test_sentinel_creates_timezone(self):
        row = ["other::stem", "*"] + [""] * (len(SIGNALS_HEADER) - 2) + ["", "", "", "", "UTC"]
        data = _build_xlsx([FULL_SIGNALS_HEADER, row])
        result = xlsx_bytes_to_database_options(data)
        assert result["other::stem"]["additional_informations"] == {"timezone": "UTC"}


TRACE_SIGNALS_HEADER = [*SIGNALS_HEADER, "trace_mode", "line_width", "opacity", "marker_symbol"]
FULL_SIGNALS_HEADER = [*TRACE_SIGNALS_HEADER, "timezone"]


class TestSentinelRowTraceOptions:
    def test_sentinel_creates_trace_options(self):
        data = _build_xlsx(
            [
                TRACE_SIGNALS_HEADER,
                [
                    "other::syringe",
                    "*",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "lines+markers",
                    "2.5",
                    "0.8",
                    "circle",
                ],
            ]
        )
        result = xlsx_bytes_to_database_options(data)
        assert result["other::syringe"]["trace_options"] == {
            "mode": "lines+markers",
            "line_width": 2.5,
            "opacity": 0.8,
            "marker_symbol": "circle",
        }

    def test_sentinel_partial_trace_options(self):
        data = _build_xlsx(
            [
                TRACE_SIGNALS_HEADER,
                [
                    "ds_a",
                    "*",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "lines+markers",
                    "",
                    "",
                    "",
                ],
            ]
        )
        result = xlsx_bytes_to_database_options(data)
        assert result["ds_a"]["trace_options"] == {"mode": "lines+markers"}

    @pytest.mark.parametrize(
        ("column_name", "column_index", "value", "expected"),
        [
            # No range/enum validation happens here -- values pass through as-is; Plotly is
            # the one that eventually rejects an out-of-range or unknown value.
            ("trace_mode", 14, "not+a+real+mode", "not+a+real+mode"),
            ("line_width", 15, "-3", -3.0),
            ("opacity", 16, "1.5", 1.5),
        ],
    )
    def test_sentinel_trace_options_are_not_range_or_enum_checked(
        self, column_name, column_index, value, expected
    ):
        row = ["ds_a", "*"] + [""] * (len(SIGNALS_HEADER) - 2) + ["", "", "", ""]
        row[column_index] = value
        data = _build_xlsx([TRACE_SIGNALS_HEADER, row])
        result = xlsx_bytes_to_database_options(data)
        key = {"trace_mode": "mode"}.get(column_name, column_name)
        assert result["ds_a"]["trace_options"] == {key: expected}

    def test_trace_mode_on_per_signal_row_is_ignored_with_warning(self, caplog):
        data = _build_xlsx(
            [
                TRACE_SIGNALS_HEADER,
                [
                    "ds_a",
                    "SIG1",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "lines+markers",
                    "",
                    "",
                    "",
                ],
            ]
        )
        result = xlsx_bytes_to_database_options(data)
        assert "trace_options" not in result["ds_a"]
        assert "trace_mode" in caplog.text

    @pytest.mark.parametrize(
        ("column_name", "value"),
        [
            ("trace_mode", "lines+markers"),
            ("line_width", "2.5"),
            ("opacity", "0.8"),
            ("marker_symbol", "circle"),
            ("timezone", "Europe/Paris"),
        ],
    )
    def test_sentinel_only_column_on_per_signal_row_warns(self, caplog, column_name, value):
        row = ["ds_a", "SIG1"] + [""] * (len(SIGNALS_HEADER) - 2) + ["", "", "", "", ""]
        row[FULL_SIGNALS_HEADER.index(column_name)] = value
        data = _build_xlsx([FULL_SIGNALS_HEADER, row])
        result = xlsx_bytes_to_database_options(data)
        assert "trace_options" not in result["ds_a"]
        assert column_name in caplog.text

    def test_trace_options_are_isolated_per_datasource(self):
        data = _build_xlsx(
            [
                TRACE_SIGNALS_HEADER,
                [
                    "ds_a",
                    "*",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "lines",
                    "",
                    "",
                    "",
                ],
                [
                    "ds_b",
                    "*",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "markers",
                    "",
                    "",
                    "",
                ],
            ]
        )
        result = xlsx_bytes_to_database_options(data)
        assert result["ds_a"]["trace_options"] == {"mode": "lines"}
        assert result["ds_b"]["trace_options"] == {"mode": "markers"}


# ---------------------------------------------------------------------------
# Tests: group scope resolution
# ---------------------------------------------------------------------------


class TestGroupResolution:
    def test_local_group_single_datasource(self):
        """Signals from one datasource → datasource grouped_fields."""
        data = _build_xlsx(
            [
                SIGNALS_HEADER,
                ["ds_a", "S1", "", "", "", "", "", "", "", "", "", "", "", "MyGroup"],
                ["ds_a", "S2", "", "", "", "", "", "", "", "", "", "", "", "MyGroup"],
            ]
        )
        result = xlsx_bytes_to_database_options(data)
        assert "grouped_fields" in result["ds_a"]
        assert result["ds_a"]["grouped_fields"]["MyGroup"] == ["S1", "S2"]
        assert "global" not in result

    def test_global_group_multiple_datasources(self):
        """Same group name across datasources → global.grouped_fields."""
        data = _build_xlsx(
            [
                SIGNALS_HEADER,
                ["ds_a", "S1", "", "", "", "", "", "", "", "", "", "", "", "SharedGroup"],
                ["ds_b", "S2", "", "", "", "", "", "", "", "", "", "", "", "SharedGroup"],
            ]
        )
        result = xlsx_bytes_to_database_options(data)
        assert "global" in result
        assert result["global"]["grouped_fields"]["SharedGroup"] == ["S1", "S2"]
        assert "grouped_fields" not in result["ds_a"]
        assert "grouped_fields" not in result["ds_b"]

    def test_mixed_local_and_global_groups(self):
        """One global group, one local group — both resolved independently."""
        data = _build_xlsx(
            [
                SIGNALS_HEADER,
                ["ds_a", "S1", "", "", "", "", "", "", "", "", "", "", "", "Global; Local"],
                ["ds_b", "S2", "", "", "", "", "", "", "", "", "", "", "", "Global"],
                ["ds_a", "S3", "", "", "", "", "", "", "", "", "", "", "", "Local"],
            ]
        )
        result = xlsx_bytes_to_database_options(data)
        # "Global" spans ds_a and ds_b → global
        assert result["global"]["grouped_fields"]["Global"] == ["S1", "S2"]
        # "Local" only in ds_a → local
        assert result["ds_a"]["grouped_fields"]["Local"] == ["S1", "S3"]

    def test_multiple_groups_semicolon_separated(self):
        data = _build_xlsx(
            [
                SIGNALS_HEADER,
                ["ds_a", "S1", "", "", "", "", "", "", "", "", "", "", "", "G1; G2"],
                ["ds_a", "S2", "", "", "", "", "", "", "", "", "", "", "", "G2"],
            ]
        )
        result = xlsx_bytes_to_database_options(data)
        assert result["ds_a"]["grouped_fields"]["G1"] == ["S1"]
        assert result["ds_a"]["grouped_fields"]["G2"] == ["S1", "S2"]


# ---------------------------------------------------------------------------
# Tests: loops sheet
# ---------------------------------------------------------------------------


class TestLoopsSheet:
    def test_loop_parsed(self):
        data = _build_xlsx(
            [SIGNALS_HEADER, ["ds_a", "S1", "", "", "", "", "", "", "", "", "", "", "", ""]],
            loops_rows=[LOOPS_HEADER, ["ds_a", "pv_loop", "Paw", "Vol"]],
        )
        result = xlsx_bytes_to_database_options(data)
        assert result["ds_a"]["loop"] == {"pv_loop": ["Paw", "Vol"]}

    def test_no_loops_sheet(self):
        data = _build_xlsx(
            [
                SIGNALS_HEADER,
                ["ds_a", "S1", "", "", "", "", "", "", "", "", "", "", "", ""],
            ]
        )
        result = xlsx_bytes_to_database_options(data)
        assert "loop" not in result["ds_a"]

    def test_loop_for_unknown_datasource_creates_entry(self):
        """A loop row whose datasource doesn't appear in signals still creates an entry."""
        data = _build_xlsx(
            [SIGNALS_HEADER, ["ds_a", "S1", "", "", "", "", "", "", "", "", "", "", "", ""]],
            loops_rows=[LOOPS_HEADER, ["ds_b", "loop1", "X", "Y"]],
        )
        result = xlsx_bytes_to_database_options(data)
        assert result["ds_b"]["loop"] == {"loop1": ["X", "Y"]}

    def test_global_loop_datasource_sentinel(self):
        """datasource='global' routes loop into result['global']['loop']."""
        data = _build_xlsx(
            [SIGNALS_HEADER, ["ds_a", "S1", "", "", "", "", "", "", "", "", "", "", "", ""]],
            loops_rows=[LOOPS_HEADER, ["global", "cross_pv", "ds_a::Paw", "ds_b::Vol"]],
        )
        result = xlsx_bytes_to_database_options(data)
        assert result["global"]["loop"] == {"cross_pv": ["ds_a::Paw", "ds_b::Vol"]}

    def test_global_loop_coexists_with_global_grouped_fields(self):
        """global.loop and global.grouped_fields can coexist under the same 'global' key."""
        data = _build_xlsx(
            [
                SIGNALS_HEADER,
                ["ds_a", "Paw", "", "", "", "", "", "", "", "", "", "", "", "MyGroup"],
                ["ds_b", "Vol", "", "", "", "", "", "", "", "", "", "", "", "MyGroup"],
            ],
            loops_rows=[LOOPS_HEADER, ["global", "pv", "ds_a::Paw", "ds_b::Vol"]],
        )
        result = xlsx_bytes_to_database_options(data)
        assert "grouped_fields" in result["global"]
        assert "loop" in result["global"]


class TestSpectrogramsSheet:
    def test_spectrogram_parsed(self):
        data = _build_xlsx(
            [SIGNALS_HEADER, ["ds_a", "S1", "", "", "", "", "", "", "", "", "", "", "", ""]],
            spectrograms_rows=[
                SPECTROGRAMS_HEADER,
                ["ds_a", "S1 spectrogram", "S1", "0.5", "30", "", ""],
            ],
        )
        result = xlsx_bytes_to_database_options(data)
        assert result["ds_a"]["spectrogram"] == {
            "S1 spectrogram": {"signal": "S1", "freq_range": [0.5, 30.0]}
        }

    def test_db_range_parsed_when_both_set(self):
        data = _build_xlsx(
            [SIGNALS_HEADER, ["ds_a", "S1", "", "", "", "", "", "", "", "", "", "", "", ""]],
            spectrograms_rows=[
                SPECTROGRAMS_HEADER,
                ["ds_a", "S1 spectrogram", "S1", "0.5", "30", "-10", "20"],
            ],
        )
        result = xlsx_bytes_to_database_options(data)
        assert result["ds_a"]["spectrogram"]["S1 spectrogram"]["db_range"] == [-10.0, 20.0]

    def test_db_range_ignored_when_only_one_bound_set(self):
        data = _build_xlsx(
            [SIGNALS_HEADER, ["ds_a", "S1", "", "", "", "", "", "", "", "", "", "", "", ""]],
            spectrograms_rows=[
                SPECTROGRAMS_HEADER,
                ["ds_a", "S1 spectrogram", "S1", "0.5", "30", "-10", ""],
            ],
        )
        result = xlsx_bytes_to_database_options(data)
        assert "db_range" not in result["ds_a"]["spectrogram"]["S1 spectrogram"]

    def test_missing_freq_range_skips_row(self):
        data = _build_xlsx(
            [SIGNALS_HEADER, ["ds_a", "S1", "", "", "", "", "", "", "", "", "", "", "", ""]],
            spectrograms_rows=[
                SPECTROGRAMS_HEADER,
                ["ds_a", "S1 spectrogram", "S1", "", "30", "", ""],
            ],
        )
        result = xlsx_bytes_to_database_options(data)
        assert "spectrogram" not in result.get("ds_a", {})

    def test_no_spectrograms_sheet(self):
        data = _build_xlsx(
            [SIGNALS_HEADER, ["ds_a", "S1", "", "", "", "", "", "", "", "", "", "", "", ""]]
        )
        result = xlsx_bytes_to_database_options(data)
        assert "spectrogram" not in result["ds_a"]

    def test_spectrogram_for_unknown_datasource_creates_entry(self):
        data = _build_xlsx(
            [SIGNALS_HEADER, ["ds_a", "S1", "", "", "", "", "", "", "", "", "", "", "", ""]],
            spectrograms_rows=[
                SPECTROGRAMS_HEADER,
                ["ds_b", "X spectrogram", "X", "0.5", "30", "", ""],
            ],
        )
        result = xlsx_bytes_to_database_options(data)
        assert result["ds_b"]["spectrogram"] == {
            "X spectrogram": {"signal": "X", "freq_range": [0.5, 30.0]}
        }

    def test_window_s_and_overlap_become_an_override(self):
        data = _build_xlsx(
            [SIGNALS_HEADER, ["ds_a", "S1", "", "", "", "", "", "", "", "", "", "", "", ""]],
            spectrograms_rows=[
                SPECTROGRAMS_HEADER,
                ["ds_a", "S1 spectrogram", "S1", "0.5", "30", "", "", "8", "0.75"],
            ],
        )
        result = xlsx_bytes_to_database_options(data)
        assert result["ds_a"]["spectrogram"]["S1 spectrogram"]["window_s"] == 8.0
        assert result["ds_a"]["spectrogram"]["S1 spectrogram"]["overlap"] == 0.75


class TestPsdsSheet:
    def test_single_signal_psd_parsed(self):
        data = _build_xlsx(
            [SIGNALS_HEADER, ["ds_a", "S1", "", "", "", "", "", "", "", "", "", "", "", ""]],
            psds_rows=[PSDS_HEADER, ["ds_a", "S1 PSD", "S1", "0.5", "30", "", ""]],
        )
        result = xlsx_bytes_to_database_options(data)
        assert result["ds_a"]["psd"] == {"S1 PSD": {"signals": ["S1"], "freq_range": [0.5, 30.0]}}

    def test_rows_sharing_a_group_overlay_on_one_plot(self):
        data = _build_xlsx(
            [SIGNALS_HEADER, ["ds_a", "S1", "", "", "", "", "", "", "", "", "", "", "", ""]],
            psds_rows=[
                PSDS_HEADER,
                ["ds_a", "EEG PSD", "S1", "0.5", "30", "", ""],
                ["ds_a", "EEG PSD", "S2", "0.5", "30", "", ""],
                ["ds_a", "EEG PSD", "S3", "", "", "", ""],
            ],
        )
        result = xlsx_bytes_to_database_options(data)
        # Later rows contribute only their signal; freq_range comes from the first row.
        assert result["ds_a"]["psd"] == {
            "EEG PSD": {"signals": ["S1", "S2", "S3"], "freq_range": [0.5, 30.0]}
        }

    def test_distinct_groups_stay_separate_plots(self):
        data = _build_xlsx(
            [SIGNALS_HEADER, ["ds_a", "S1", "", "", "", "", "", "", "", "", "", "", "", ""]],
            psds_rows=[
                PSDS_HEADER,
                ["ds_a", "Low band", "S1", "0.04", "0.4", "", ""],
                ["ds_a", "High band", "S1", "0.5", "30", "", ""],
            ],
        )
        result = xlsx_bytes_to_database_options(data)
        assert sorted(result["ds_a"]["psd"]) == ["High band", "Low band"]
        assert result["ds_a"]["psd"]["Low band"]["freq_range"] == [0.04, 0.4]

    def test_row_with_no_groups_is_excluded_from_every_plot(self):
        data = _build_xlsx(
            [SIGNALS_HEADER, ["ds_a", "S1", "", "", "", "", "", "", "", "", "", "", "", ""]],
            psds_rows=[PSDS_HEADER, ["ds_a", "", "S1", "0.5", "30", "", ""]],
        )
        result = xlsx_bytes_to_database_options(data)
        assert "psd" not in result.get("ds_a", {})

    def test_row_with_multiple_groups_becomes_a_trace_on_each_plot(self):
        data = _build_xlsx(
            [SIGNALS_HEADER, ["ds_a", "S1", "", "", "", "", "", "", "", "", "", "", "", ""]],
            psds_rows=[
                PSDS_HEADER,
                ["ds_a", "Low band;High band", "S1", "0.04", "30", "", ""],
                ["ds_a", "High band", "S2", "0.04", "30", "", ""],
            ],
        )
        result = xlsx_bytes_to_database_options(data)
        assert result["ds_a"]["psd"]["Low band"]["signals"] == ["S1"]
        assert result["ds_a"]["psd"]["High band"]["signals"] == ["S1", "S2"]

    def test_conflicting_freq_range_keeps_the_first_and_warns(self, caplog):
        data = _build_xlsx(
            [SIGNALS_HEADER, ["ds_a", "S1", "", "", "", "", "", "", "", "", "", "", "", ""]],
            psds_rows=[
                PSDS_HEADER,
                ["ds_a", "EEG PSD", "S1", "0.5", "30", "", ""],
                ["ds_a", "EEG PSD", "S2", "1", "40", "", ""],
            ],
        )
        with caplog.at_level("WARNING"):
            result = xlsx_bytes_to_database_options(data)
        assert result["ds_a"]["psd"]["EEG PSD"]["freq_range"] == [0.5, 30.0]
        assert any(
            "freq_range" in message and "conflicts" in message for message in caplog.messages
        )

    def test_conflicting_db_range_keeps_the_first_and_warns(self, caplog):
        data = _build_xlsx(
            [SIGNALS_HEADER, ["ds_a", "S1", "", "", "", "", "", "", "", "", "", "", "", ""]],
            psds_rows=[
                PSDS_HEADER,
                ["ds_a", "EEG PSD", "S1", "0.5", "30", "40", "90"],
                ["ds_a", "EEG PSD", "S2", "0.5", "30", "0", "50"],
            ],
        )
        with caplog.at_level("WARNING"):
            result = xlsx_bytes_to_database_options(data)
        assert result["ds_a"]["psd"]["EEG PSD"]["db_range"] == [40.0, 90.0]
        assert any("db_range" in message and "conflicts" in message for message in caplog.messages)

    def test_window_s_and_overlap_become_a_per_entry_override(self):
        data = _build_xlsx(
            [SIGNALS_HEADER, ["ds_a", "S1", "", "", "", "", "", "", "", "", "", "", "", ""]],
            psds_rows=[
                PSDS_HEADER,
                ["ds_a", "EEG PSD", "S1", "0.5", "30", "", "", "2", "0.5", ""],
                ["ds_a", "EEG PSD", "S1", "0.5", "30", "", "", "8", "0.5", "wide window"],
            ],
        )
        result = xlsx_bytes_to_database_options(data)
        signals = result["ds_a"]["psd"]["EEG PSD"]["signals"]
        assert signals[0] == {"signal": "S1", "window_s": 2.0, "overlap": 0.5}
        assert signals[1] == {
            "signal": "S1",
            "window_s": 8.0,
            "overlap": 0.5,
            "label": "wide window",
        }

    def test_color_and_line_dash_become_a_per_entry_override(self):
        data = _build_xlsx(
            [SIGNALS_HEADER, ["ds_a", "S1", "", "", "", "", "", "", "", "", "", "", "", ""]],
            psds_rows=[
                PSDS_HEADER,
                ["ds_a", "EEG PSD", "S1", "0.5", "30", "", "", "", "", "a", "red", "dash"],
                ["ds_a", "EEG PSD", "S1", "0.5", "30", "", "", "", "", "b", "blue", "dot"],
            ],
        )
        result = xlsx_bytes_to_database_options(data)
        signals = result["ds_a"]["psd"]["EEG PSD"]["signals"]
        assert signals[0] == {"signal": "S1", "label": "a", "color": "red", "line_dash": "dash"}
        assert signals[1] == {"signal": "S1", "label": "b", "color": "blue", "line_dash": "dot"}

    def test_signal_without_overrides_stays_a_plain_string(self):
        data = _build_xlsx(
            [SIGNALS_HEADER, ["ds_a", "S1", "", "", "", "", "", "", "", "", "", "", "", ""]],
            psds_rows=[PSDS_HEADER, ["ds_a", "EEG PSD", "S1", "0.5", "30", "", ""]],
        )
        result = xlsx_bytes_to_database_options(data)
        assert result["ds_a"]["psd"]["EEG PSD"]["signals"] == ["S1"]

    def test_db_range_parsed_when_both_set(self):
        data = _build_xlsx(
            [SIGNALS_HEADER, ["ds_a", "S1", "", "", "", "", "", "", "", "", "", "", "", ""]],
            psds_rows=[PSDS_HEADER, ["ds_a", "S1 PSD", "S1", "0.5", "30", "40", "90"]],
        )
        result = xlsx_bytes_to_database_options(data)
        assert result["ds_a"]["psd"]["S1 PSD"]["db_range"] == [40.0, 90.0]

    def test_db_range_ignored_when_only_one_bound_set(self):
        data = _build_xlsx(
            [SIGNALS_HEADER, ["ds_a", "S1", "", "", "", "", "", "", "", "", "", "", "", ""]],
            psds_rows=[PSDS_HEADER, ["ds_a", "S1 PSD", "S1", "0.5", "30", "40", ""]],
        )
        result = xlsx_bytes_to_database_options(data)
        assert "db_range" not in result["ds_a"]["psd"]["S1 PSD"]

    def test_missing_freq_range_skips_row(self):
        data = _build_xlsx(
            [SIGNALS_HEADER, ["ds_a", "S1", "", "", "", "", "", "", "", "", "", "", "", ""]],
            psds_rows=[PSDS_HEADER, ["ds_a", "S1 PSD", "S1", "", "30", "", ""]],
        )
        result = xlsx_bytes_to_database_options(data)
        assert "psd" not in result.get("ds_a", {})

    def test_no_psds_sheet(self):
        data = _build_xlsx(
            [SIGNALS_HEADER, ["ds_a", "S1", "", "", "", "", "", "", "", "", "", "", "", ""]]
        )
        result = xlsx_bytes_to_database_options(data)
        assert "psd" not in result["ds_a"]


# ---------------------------------------------------------------------------
# Tests: error handling
# ---------------------------------------------------------------------------


class TestErrorHandling:
    def test_missing_signals_sheet_raises(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "other"
        buf = io.BytesIO()
        wb.save(buf)
        with pytest.raises(ValueError, match="signals"):
            xlsx_bytes_to_database_options(buf.getvalue())

    def test_missing_required_columns_raises(self):
        data = _build_xlsx([["only_col"], ["val"]])
        with pytest.raises(ValueError, match="missing required columns"):
            xlsx_bytes_to_database_options(data)

    def test_empty_rows_ignored(self):
        data = _build_xlsx(
            [
                SIGNALS_HEADER,
                ["", "", "", "", "", "", "", "", "", "", "", "", "", ""],
                ["ds_a", "SIG1", "", "", "", "", "", "", "", "", "", "", "", ""],
            ]
        )
        result = xlsx_bytes_to_database_options(data)
        assert list(result.keys()) == ["ds_a"]


# ---------------------------------------------------------------------------
# Tests: round-trip with example file
# ---------------------------------------------------------------------------


class TestExampleFileRoundTrip:
    """Smoke test against the shipped example XLSX."""

    def test_example_xlsx_parses_without_error(self):

        example = Path("example/option_files/example_database_options.xlsx")
        if not example.exists():
            pytest.skip("Example XLSX not found")
        data = example.read_bytes()
        result = xlsx_bytes_to_database_options(data)
        assert "other::waves" in result
        assert "eit" in result
        assert "global" in result

    def test_example_passes_validation(self):
        from clinical_scope.database_options_parser import validate_database_options

        example = Path("example/option_files/example_database_options.xlsx")
        if not example.exists():
            pytest.skip("Example XLSX not found")
        result = xlsx_bytes_to_database_options(example.read_bytes())
        issues = validate_database_options(result)
        assert not any(i.severity in ("error", "warning") for i in issues)
